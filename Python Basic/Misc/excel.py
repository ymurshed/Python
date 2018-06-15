# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

file = 'D:/Video Tutorials/data.xlsx'

# Load in the workbook
wb = load_workbook(file, data_only=True)

# Get sheet names
print(wb.get_sheet_names())

# Get a sheet by name 
sheet = wb.get_sheet_by_name('Input')

# Print the sheet title 
sheet.title

# Get currently active sheet
anotherSheet = wb.active


for cellObj in sheet['A1':'C5']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)
      print('--- END ---')
      
df = pd.DataFrame(sheet.values)
print(df)
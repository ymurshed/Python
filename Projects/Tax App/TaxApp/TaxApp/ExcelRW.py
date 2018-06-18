import openpyxl, pprint
import numpy as np
from Constants import Constants as C
from ExcelDS import Entity, Post_86_Passive, Post_86_General, Post_86_Other, Pre_87_Passive, Pre_87_General, Pre_87_Other
from FileHelper import FileHelper

class ExcelRW(object):
    """description of class"""

    file = ''
    
    def __load_workbook(self):
        wb = openpyxl.load_workbook(self.file)
        return wb

    def __get_sheet(self, wb, sheet_name):
        sheet = wb.get_sheet_by_name(sheet_name)
        return sheet

    def __get_sheet_companies(self, wb, sheet_name):
        try:
            sheet = self.__get_sheet(wb, sheet_name)

            company_list = []
            cell_name = ''
            cell_index = 0
            
            if sheet_name == C.INPUT_SHEET1_NAME:
                cell_name = C.COMPANY_SHEET1_START_CELL_NAME
                cell_index = C.COMPANY_SHEET1_START_CELL_INDEX
            
            elif sheet_name == C.INPUT_SHEET2_NAME:
                cell_name = C.COMPANY_SHEET2_START_CELL_NAME
                cell_index = C.COMPANY_SHEET2_START_CELL_INDEX

            for row in range(cell_index, sheet.max_row - 1):
                item = str(sheet[cell_name + str(row)].value).strip()

                if item not in company_list:
                    company_list.append(item)
            
            return company_list

        except Exception as e:
            print(str(e))

    def __get_entity_name(self, wb, sheet_name):
        try:
            sheet = self.__get_sheet(wb, sheet_name)

            entity_dic = {}

            for row in range(13, sheet.max_row - 1):
                _key = str(sheet['C' + str(row)].value).strip()
                _value = str(sheet['D' + str(row)].value).strip()

                if _key not in entity_dic:
                    entity_dic[_key] = _value
            
            return entity_dic

        except Exception as e:
            print(str(e))

    def __get_full_data_dic(self, wb, sheet_name, companies, des, ep_tax_name, type):
        try:
            sheet = self.__get_sheet(wb, sheet_name)
            start = 18
            com_cell = 'B'
            desc_cell = 'D'
            ep_tax_cell = 'E'
            
            if sheet_name == C.INPUT_SHEET2_NAME:
                start = 13
                com_cell = 'C'
                desc_cell = 'J'
            
            dic = {}

            for l in companies:
                _key = l
                _value = 0
                
                for row in range(start, sheet.max_row - 1):
                    
                    # if company name not matched with current cell company, then continue 
                    com = str(sheet[com_cell + str(row)].value).strip()
                    if l != com:
                        continue
                        
                    descrp = str(sheet[desc_cell + str(row)].value).strip()
                    if des == C.DESCRIPTION_OTHER and (descrp == C.DESCRIPTION_GENERAL or descrp == C.DESCRIPTION_PASSIVE):
                        continue
                    elif descrp != des:
                        continue

                    if sheet_name == C.INPUT_SHEET1_NAME:
                        ep_tax = str(sheet[ep_tax_cell + str(row)].value).strip().lower()
                        if ep_tax != ep_tax_name:
                            continue

                    return_value = self.__do_calculation(sheet, type, row, sheet_name, ep_tax_name)
                    if return_value is not None:
                        _value += return_value
                
                dic[_key] = _value
            
            return dic

        except Exception as e:
            print(str(e))

    def __do_calculation(self, sheet, type, row, sheet_name, ep_tax_name):
        if type is C.EP_TYPE_C3 or type is C.TAX_TYPE_C3:
            if sheet_name == C.INPUT_SHEET2_NAME:
                if ep_tax_name == 'ep':
                    return (sheet['M13' + str(row)].value)
                else:
                    return (sheet['013' + str(row)].value)
            else:
                return (sheet['I' + str(row)].value)

        elif type is C.EP_TYPE_C2 or type is C.TAX_TYPE_C2:
            if sheet_name == C.INPUT_SHEET2_NAME:
                if ep_tax_name == 'ep':
                    return (sheet['M14' + str(row)].value)
                else:
                    return 0
            else:
                return (sheet['L' + str(row)].value + sheet['M' + str(row)].value)

        elif type is C.EP_TYPE_C1 or type is C.TAX_TYPE_C1:
            if sheet_name == C.INPUT_SHEET2_NAME:
                if ep_tax_name == 'ep':
                    return (sheet['M15' + str(row)].value)
                else:
                    return 0
            else:
                return (sheet['N' + str(row)].value + sheet['O' + str(row)].value +  
                    sheet['P' + str(row)].value + sheet['Q' + str(row)].value)

    def get_result(self, filepath):
        try:
            self.file = filepath
            wb = self.__load_workbook()

            # get Entity object
            ent = Entity()
            ent.EntityCode = list(set().union(self.__get_sheet_companies(wb, C.INPUT_SHEET1_NAME), 
                                              self.__get_sheet_companies(wb, C.INPUT_SHEET2_NAME)))
            ent.EntityCode.sort()
            ent.Name = self.__get_entity_name(wb, C.INPUT_SHEET2_NAME)
            print(ent.EntityCode)

            '''86'''
            # 86 General
            _post_86_General = Post_86_General()
            _post_86_General.dic_Post_86_General_EP_959c3 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_EP, C.EP_TYPE_C3)
            _post_86_General.dic_Post_86_General_EP_959c2 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_EP, C.EP_TYPE_C2)
            _post_86_General.dic_Post_86_General_EP_959c1 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_EP, C.EP_TYPE_C1)

            _post_86_General.dic_Post_86_General_TAX_959c3Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C3)
            _post_86_General.dic_Post_86_General_TAX_959c2Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C2)
            _post_86_General.dic_Post_86_General_TAX_959c1Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C1)

            # 86 Passive
            _post_86_Passive = Post_86_Passive()
            _post_86_Passive.dic_Post_86_Passive_EP_959c3 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_EP, C.EP_TYPE_C3)
            _post_86_Passive.dic_Post_86_Passive_EP_959c2 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_EP, C.EP_TYPE_C2)
            _post_86_Passive.dic_Post_86_Passive_EP_959c1 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_EP, C.EP_TYPE_C1)

            _post_86_Passive.dic_Post_86_Passive_TAX_959c3Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C3)
            _post_86_Passive.dic_Post_86_Passive_TAX_959c2Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C2)
            _post_86_Passive.dic_Post_86_Passive_TAX_959c1Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C1)

            # 86 Other
            _post_86_Other = Post_86_Other()
            _post_86_Other.dic_Post_86_Other_EP_959c3 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_EP, C.EP_TYPE_C3)
            _post_86_Other.dic_Post_86_Other_EP_959c2 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_EP, C.EP_TYPE_C2)
            _post_86_Other.dic_Post_86_Other_EP_959c1 = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_EP, C.EP_TYPE_C1)

            _post_86_Other.dic_Post_86_Other_TAX_959c3Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C3)
            _post_86_Other.dic_Post_86_Other_TAX_959c2Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C2)
            _post_86_Other.dic_Post_86_Other_TAX_959c1Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET1_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C1)
            
            '''87'''
            # 87 General
            _pre_87_General = Pre_87_General()
            _pre_87_General.dic_Pre_87_General_EP_959c3 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_EP, C.EP_TYPE_C3)
            _pre_87_General.dic_Pre_87_General_EP_959c2 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_EP, C.EP_TYPE_C2)
            _pre_87_General.dic_Pre_87_General_EP_959c1 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_EP, C.EP_TYPE_C1)

            _pre_87_General.dic_Pre_87_General_TAX_959c3Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C3)
            _pre_87_General.dic_Pre_87_General_TAX_959c2Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C2)
            _pre_87_General.dic_Pre_87_General_TAX_959c1Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_GENERAL, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C1)

            # 87 Passive
            _pre_87_Passive = Pre_87_Passive()
            _pre_87_Passive.dic_Pre_87_Passive_EP_959c3 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_EP, C.EP_TYPE_C3)
            _pre_87_Passive.dic_Pre_87_Passive_EP_959c2 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_EP, C.EP_TYPE_C2)
            _pre_87_Passive.dic_Pre_87_Passive_EP_959c1 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_EP, C.EP_TYPE_C1)

            _pre_87_Passive.dic_Pre_87_Passive_TAX_959c3Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C3)
            _pre_87_Passive.dic_Pre_87_Passive_TAX_959c2Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C2)
            _pre_87_Passive.dic_Pre_87_Passive_TAX_959c1Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_PASSIVE, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C1)

            # 87 Other
            _pre_87_Other = Pre_87_Other()
            _pre_87_Other.dic_Pre_87_Other_EP_959c3 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_EP, C.EP_TYPE_C3)
            _pre_87_Other.dic_Pre_87_Other_EP_959c2 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_EP, C.EP_TYPE_C2)
            _pre_87_Other.dic_Pre_87_Other_EP_959c1 = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_EP, C.EP_TYPE_C1)

            _pre_87_Other.dic_Pre_87_Other_TAX_959c3Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C3)
            _pre_87_Other.dic_Pre_87_Other_TAX_959c2Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C2)
            _pre_87_Other.dic_Pre_87_Other_TAX_959c1Tax = self.__get_full_data_dic(wb, C.INPUT_SHEET2_NAME, ent.EntityCode, C.DESCRIPTION_OTHER, C.EP_TAX_NAME_TAX, C.TAX_TYPE_C1)
            
            self.__set_sheet_data(ent, _post_86_General, _post_86_Passive, _post_86_Other,
                                  _pre_87_General, _pre_87_Passive, _pre_87_Other)

            
        except Exception as e:
            print(str(e))

    def __set_sheet_data(self, ent, g86, p86, o86, g87, p87, o87):
        try:
            fh = FileHelper()
            filepath = fh.get_file_path(C.OUTPUT_FILE)
            
            self.file = filepath
            wb = self.__load_workbook()
            sheet = self.__get_sheet(wb, 'Pools')

            i = 7
            for el in ent.EntityCode:
                sheet['A' + str(i)].value = el
                
                if el in ent.Name.keys():
                    sheet['B' + str(i)].value = ent.Name[el]           
                
                # 86 General
                if el in g86.dic_Post_86_General_EP_959c3.keys():
                    sheet['D' + str(i)].value = g86.dic_Post_86_General_EP_959c3[el]
                else:
                    sheet['D' + str(i)].value = 0

                if el in g86.dic_Post_86_General_EP_959c2.keys():
                    sheet['E' + str(i)].value = g86.dic_Post_86_General_EP_959c2[el]
                else:
                    sheet['E' + str(i)].value = 0

                if el in g86.dic_Post_86_General_EP_959c1.keys():
                    sheet['F' + str(i)].value = g86.dic_Post_86_General_EP_959c1[el]
                else:
                    sheet['F' + str(i)].value = 0

                if el in g86.dic_Post_86_General_TAX_959c3Tax.keys():
                    sheet['G' + str(i)].value = g86.dic_Post_86_General_TAX_959c3Tax[el]
                else:
                    sheet['G' + str(i)].value = 0

                if el in g86.dic_Post_86_General_TAX_959c2Tax.keys():
                    sheet['H' + str(i)].value = g86.dic_Post_86_General_TAX_959c2Tax[el]
                else:
                    sheet['H' + str(i)].value = 0

                if el in g86.dic_Post_86_General_TAX_959c1Tax.keys():
                    sheet['I' + str(i)].value = g86.dic_Post_86_General_TAX_959c1Tax[el]
                else:
                    sheet['I' + str(i)].value = 0
                
                # 86 Passive
                if el in p86.dic_Post_86_Passive_EP_959c3.keys():
                    sheet['K' + str(i)].value = p86.dic_Post_86_Passive_EP_959c3[el]
                else:
                    sheet['K' + str(i)].value = 0

                if el in p86.dic_Post_86_Passive_EP_959c2.keys():
                    sheet['L' + str(i)].value = p86.dic_Post_86_Passive_EP_959c2[el]
                else:
                    sheet['L' + str(i)].value = 0

                if el in p86.dic_Post_86_Passive_EP_959c1.keys():
                    sheet['M' + str(i)].value = p86.dic_Post_86_Passive_EP_959c1[el]
                else:
                    sheet['M' + str(i)].value = 0

                if el in p86.dic_Post_86_Passive_TAX_959c3Tax.keys():
                    sheet['N' + str(i)].value = p86.dic_Post_86_Passive_TAX_959c3Tax[el]
                else:
                    sheet['N' + str(i)].value = 0

                if el in p86.dic_Post_86_Passive_TAX_959c2Tax.keys():
                    sheet['O' + str(i)].value = p86.dic_Post_86_Passive_TAX_959c2Tax[el]
                else:
                    sheet['O' + str(i)].value = 0

                if el in p86.dic_Post_86_Passive_TAX_959c1Tax.keys():
                    sheet['P' + str(i)].value = p86.dic_Post_86_Passive_TAX_959c1Tax[el]
                else:
                    sheet['P' + str(i)].value = 0

                # 86 Other
                if el in o86.dic_Post_86_Other_EP_959c3.keys():
                    sheet['R' + str(i)].value = o86.dic_Post_86_Other_EP_959c3[el]
                else:
                    sheet['R' + str(i)].value = 0

                if el in o86.dic_Post_86_Other_EP_959c2.keys():
                    sheet['S' + str(i)].value = o86.dic_Post_86_Other_EP_959c2[el]
                else:
                    sheet['S' + str(i)].value = 0

                if el in o86.dic_Post_86_Other_EP_959c1.keys():
                    sheet['T' + str(i)].value = o86.dic_Post_86_Other_EP_959c1[el]
                else:
                    sheet['T' + str(i)].value = 0

                if el in o86.dic_Post_86_Other_TAX_959c3Tax.keys():
                    sheet['U' + str(i)].value = o86.dic_Post_86_Other_TAX_959c3Tax[el]
                else:
                    sheet['U' + str(i)].value = 0

                if el in o86.dic_Post_86_Other_TAX_959c2Tax.keys():
                    sheet['V' + str(i)].value = o86.dic_Post_86_Other_TAX_959c2Tax[el]
                else:
                    sheet['V' + str(i)].value = 0

                if el in o86.dic_Post_86_Other_TAX_959c1Tax.keys():
                    sheet['W' + str(i)].value = o86.dic_Post_86_Other_TAX_959c1Tax[el]
                else:
                    sheet['W' + str(i)].value = 0

                i += 1
                #TODO: Save data for the next tab

            wb.save(self.file)

        except Exception as e:
            print(str(e))

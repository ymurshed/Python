import openpyxl, pprint
import numpy as np

class ExceIO(object):
    """description of class"""

    file = ''
    def __load_workbook(self):
        wb = openpyxl.load_workbook(self.file)
        return wb

    def __get_sheet(self, wb, sheet_name):
        sheet = wb.get_sheet_by_name(sheet_name)
        return sheet

    def __get_sheet_data(self):
        try:
            wb = self.__load_workbook()
            a = None
            
            j = 0
            for sheet_no in range(1, 3):
                sheet_name = 'Sheet' + str(sheet_no)
                sheet = self.__get_sheet(wb, sheet_name)

                #construct np 3d array
                if a is None:
                    row = (sheet.max_row + 1) - 2
                    col = 3
                    a = np.arange(2 * row * col).reshape(2, row, col)
                
                i = 0
                for row in range(2, sheet.max_row + 1):
                    col1 = sheet['A' + str(row)].value
                    col2 = sheet['B' + str(row)].value
                    col3 = sheet['C' + str(row)].value
                    l = [col1, col2, col3]
                    a[j][i] = l
                    i += 1
                
                j += 1
                print(a)

            return a

        except Exception as e:
            print(str(e))

    def __set_sheet_data(self, result):
        try:
            wb = self.__load_workbook()
            wb.create_sheet(index=2, title='Result')
            sheet = self.__get_sheet(wb, 'Result')

            #setting up results
            sheet['A1'] = 'Col1'
            sheet['B1'] = 'Col2'
            sheet['C1'] = 'Col3'

            i = 0
            j = 2
            for element in result.flat:
                if i == 0:
                    sheet['A' + str(j)].value = element
                elif i == 1:
                    sheet['B' + str(j)].value = element
                else:
                   sheet['C' + str(j)].value = element
                
                i += 1
                if i == 3:
                    i = 0
                    j += 1
            
            wb.save(self.file)

        except Exception as e:
            print(str(e))

    def __do_operation(self, operationType, a):
        return {
            'add': a[1] + a[0],
            'sub': a[1] - a[0],
            'div': a[1] / a[0],
            'mul': a[1] * a[0],
        }.get(operationType, a[1] + a[0])

    def get_result(self, filepath, operationType):
        self.file = filepath
        a = self.__get_sheet_data()
        result = self.__do_operation(operationType, a)
        self.__set_sheet_data(result)
        print(result)

    